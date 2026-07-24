from __future__ import annotations

import csv
import io
import json
import uuid
from datetime import UTC, date, datetime
from decimal import Decimal
from typing import Any, Iterable

from django.db.models import Q
from django.http import HttpResponse
from openpyxl import Workbook
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.audit.models import AuditEvent
from apps.audit.services import AuditService
from apps.payouts.models import BalanceEntry, PayoutRequest
from apps.payouts.ops_selectors import (
    build_payout_admin_ops_summary,
    build_payout_integrity_snapshot,
    build_payout_repair_preview,
)
from apps.payouts.services import PayoutService
from common.csv_safe import csv_safe_value, spreadsheet_safe_value

EXPORT_LIMIT = 10_000


def _query_value(params, key: str) -> str:
    value = params.get(key, "")
    return str(value).strip() if value is not None else ""


def _positive_int(value: str, *, default: int, maximum: int) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return max(1, min(parsed, maximum))


def _money(value: Decimal | None) -> str:
    amount = value if value is not None else Decimal("0.00")
    return f"{amount.quantize(Decimal('0.01'))}"


def _apply_created_filters(queryset, params):
    created_from = _query_value(params, "created_from")
    created_to = _query_value(params, "created_to")
    if created_from:
        queryset = queryset.filter(created_at__gte=created_from)
    if created_to:
        queryset = queryset.filter(created_at__lte=created_to)
    return queryset


def _audit_context(*, request, export_type: str, filename: str, filters: dict, exported_rows: int, total_rows: int, limit: int) -> None:
    AuditService.log_admin_action(
        request=request,
        action="payouts.admin_ops.csv_export",
        target_type="payout_export",
        target_id=export_type,
        context={
            "export_type": export_type,
            "filename": filename,
            "filters": filters,
            "exported_rows": exported_rows,
            "total_rows": total_rows,
            "limit": limit,
            "truncated": total_rows > exported_rows,
        },
    )




def _json_safe(value: Any):
    if isinstance(value, Decimal):
        return _money(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, dict):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_json_safe(item) for item in value]
    return value

def _csv_response(*, filename: str, header: list[str], rows: Iterable[list[str]]) -> HttpResponse:
    buffer = io.StringIO()
    buffer.write("\ufeff")
    writer = csv.writer(buffer)
    writer.writerow(header)
    writer.writerows([[csv_safe_value(value) for value in row] for row in rows])

    response = HttpResponse(buffer.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _xlsx_response(*, filename: str, workbook: Workbook) -> HttpResponse:
    buffer = io.BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _xlsx_safe_row(values: Iterable[Any]) -> list[Any]:
    return [spreadsheet_safe_value(value) for value in values]


def _filters_from_params(params) -> dict[str, str]:
    return {key: _query_value(params, key) for key in ["status", "trainer_id", "currency", "created_from", "created_to"]}


def _payout_queryset_for_export(params):
    queryset = PayoutRequest.objects.select_related("trainer", "trainer__user", "wallet").order_by("-created_at")
    status_filter = _query_value(params, "status")
    trainer_id = _query_value(params, "trainer_id")
    currency = _query_value(params, "currency")
    if status_filter:
        if status_filter == PayoutRequest.Status.PENDING:
            queryset = queryset.filter(status__in=[PayoutRequest.Status.PENDING, PayoutRequest.Status.REQUESTED])
        else:
            queryset = queryset.filter(status=status_filter)
    if trainer_id:
        queryset = queryset.filter(Q(trainer__user_id=trainer_id) | Q(trainer_id=trainer_id))
    if currency:
        queryset = queryset.filter(currency=currency.upper())
    return _apply_created_filters(queryset, params)


def _ledger_queryset_for_export(params, *, entry_type: str):
    queryset = (
        BalanceEntry.objects.select_related("wallet", "wallet__trainer", "wallet__trainer__user")
        .filter(entry_type=entry_type)
        .order_by("-created_at")
    )
    trainer_id = _query_value(params, "trainer_id")
    currency = _query_value(params, "currency")
    if trainer_id:
        queryset = queryset.filter(Q(wallet__trainer__user_id=trainer_id) | Q(wallet__trainer_id=trainer_id))
    if currency:
        queryset = queryset.filter(currency=currency.upper())
    return _apply_created_filters(queryset, params)


def _balance_queryset_for_export(params):
    from apps.payouts.models import TrainerWallet

    queryset = TrainerWallet.objects.select_related("trainer", "trainer__user").order_by("trainer_id", "currency", "id")
    trainer_id = _query_value(params, "trainer_id")
    currency = _query_value(params, "currency")
    if trainer_id:
        queryset = queryset.filter(Q(trainer__user_id=trainer_id) | Q(trainer_id=trainer_id))
    if currency:
        queryset = queryset.filter(currency=currency.upper())
    return queryset


RECONCILIATION_EXPORT_HEADER = [
    "section",
    "id",
    "trainer_id",
    "trainer_name",
    "wallet_id",
    "payout_id",
    "entry_type",
    "status",
    "direction",
    "amount",
    "currency",
    "source_type",
    "source_id",
    "available_amount",
    "pending_amount",
    "locked_amount",
    "integrity_status",
    "issue_code",
    "severity",
    "message",
    "created_at",
    "updated_at",
]


def _blank_export_row(section: str) -> dict[str, str]:
    return {key: "" for key in RECONCILIATION_EXPORT_HEADER} | {"section": section}


def _payout_export_rows(payouts) -> list[dict[str, str]]:
    rows = []
    for payout in payouts:
        row = _blank_export_row("payout")
        row.update(
            {
                "id": str(payout.id),
                "trainer_id": str(payout.trainer.user_id),
                "trainer_name": getattr(payout.trainer, "display_name", ""),
                "wallet_id": str(payout.wallet_id),
                "payout_id": str(payout.id),
                "status": payout.status,
                "amount": _money(payout.amount),
                "currency": payout.currency,
                "created_at": payout.created_at.isoformat(),
                "updated_at": payout.updated_at.isoformat(),
            }
        )
        rows.append(row)
    return rows


def _ledger_export_rows(entries, *, section: str) -> list[dict[str, str]]:
    rows = []
    for entry in entries:
        payout_id = str(entry.source_id) if entry.source_type == "payout_request" else ""
        row = _blank_export_row(section)
        row.update(
            {
                "id": str(entry.id),
                "trainer_id": str(entry.wallet.trainer.user_id),
                "trainer_name": getattr(entry.wallet.trainer, "display_name", ""),
                "wallet_id": str(entry.wallet_id),
                "payout_id": payout_id,
                "entry_type": entry.entry_type,
                "status": entry.status,
                "direction": entry.direction,
                "amount": _money(entry.amount),
                "currency": entry.currency,
                "source_type": entry.source_type,
                "source_id": str(entry.source_id),
                "created_at": entry.created_at.isoformat(),
                "updated_at": entry.updated_at.isoformat(),
            }
        )
        rows.append(row)
    return rows


def _balance_export_rows(wallets) -> list[dict[str, str]]:
    rows = []
    for wallet in wallets:
        row = _blank_export_row("balance")
        row.update(
            {
                "id": str(wallet.id),
                "trainer_id": str(wallet.trainer.user_id),
                "trainer_name": getattr(wallet.trainer, "display_name", ""),
                "wallet_id": str(wallet.id),
                "currency": wallet.currency,
                "available_amount": _money(wallet.available_amount),
                "pending_amount": _money(wallet.pending_amount),
                "locked_amount": _money(wallet.locked_amount),
                "created_at": wallet.created_at.isoformat(),
                "updated_at": wallet.updated_at.isoformat(),
            }
        )
        rows.append(row)
    return rows


def _integrity_export_rows(snapshot: dict[str, Any]) -> list[dict[str, str]]:
    summary = snapshot.get("summary", {})
    rows = []
    summary_row = _blank_export_row("integrity_status")
    summary_row.update(
        {
            "integrity_status": str(summary.get("status", "unknown")),
            "amount": str(summary.get("issue_count", 0)),
            "message": (
                f"issues={summary.get('issue_count', 0)}; wallets={summary.get('wallet_count', 0)}; "
                f"payouts={summary.get('payouts_scanned', 0)}; ledger={summary.get('ledger_entries_scanned', 0)}"
            ),
            "created_at": str(snapshot.get("generated_at", "")),
        }
    )
    rows.append(summary_row)
    for issue in snapshot.get("issues", []):
        row = _blank_export_row("integrity_issue")
        row.update(
            {
                "id": str(issue.get("ledger_entry_id") or issue.get("payout_id") or issue.get("wallet_id") or ""),
                "trainer_id": str(issue.get("trainer_id") or ""),
                "wallet_id": str(issue.get("wallet_id") or ""),
                "payout_id": str(issue.get("payout_id") or ""),
                "status": str(issue.get("status") or ""),
                "amount": str(issue.get("amount") or issue.get("payout_amount") or issue.get("reserve_amount") or issue.get("delta") or ""),
                "currency": str(issue.get("currency") or issue.get("payout_currency") or issue.get("wallet_currency") or ""),
                "integrity_status": str(summary.get("status", "unknown")),
                "issue_code": str(issue.get("code") or ""),
                "severity": str(issue.get("severity") or ""),
                "message": str(issue.get("message") or ""),
            }
        )
        rows.append(row)
    return rows


def _build_reconciliation_export_dataset(params, *, limit: int) -> dict[str, Any]:
    filters = _filters_from_params(params)
    snapshot_params = {**filters, "limit": max(limit, 100)}
    payouts = list(_payout_queryset_for_export(params)[:limit])
    reserves = list(_ledger_queryset_for_export(params, entry_type=BalanceEntry.EntryType.RESERVE)[:limit])
    releases = list(_ledger_queryset_for_export(params, entry_type=BalanceEntry.EntryType.RELEASE)[:limit])
    balances = list(_balance_queryset_for_export(params)[:limit])
    integrity = build_payout_integrity_snapshot(snapshot_params)
    sections = {
        "payouts": _payout_export_rows(payouts),
        "reserves": _ledger_export_rows(reserves, section="reserve"),
        "releases": _ledger_export_rows(releases, section="release"),
        "balances": _balance_export_rows(balances),
        "integrity_status": _integrity_export_rows(integrity),
    }
    return {
        "filters": filters,
        "sections": sections,
        "integrity": integrity,
        "counts": {name: len(rows) for name, rows in sections.items()},
    }


def _audit_reconciliation_export(*, request, export_format: str, filename: str, filters: dict[str, str], counts: dict[str, int], limit: int) -> None:
    AuditService.log_admin_action(
        request=request,
        action="payouts.reconciliation_report_export",
        target_type="payout_reconciliation_export",
        target_id=export_format,
        context={
            "export_type": "reconciliation_report",
            "format": export_format,
            "filename": filename,
            "filters": filters,
            "section_counts": counts,
            "exported_rows": sum(counts.values()),
            "limit": limit,
        },
    )


REPAIR_AUDIT_EXPORT_HEADER = [
    "repair_id",
    "operator_id",
    "operator_email",
    "timestamp",
    "action",
    "status",
    "batch_size",
    "repaired_count",
    "skipped_count",
    "manual_review_count",
    "actions",
    "result",
    "filters",
]

REPAIR_AUDIT_ACTION_HEADER = [
    "repair_id",
    "timestamp",
    "operator_email",
    "issue_code",
    "action_code",
    "status",
    "payout_id",
    "wallet_id",
    "ledger_entry_id",
    "amount",
    "currency",
    "reason",
]


def _repair_audit_queryset(params):
    queryset = (
        AuditEvent.objects.select_related("actor")
        .filter(event_type="admin.payouts.repair_execution", entity_type="payout_repair_execution")
        .order_by("-created_at", "-id")
    )
    created_from = _query_value(params, "created_from")
    created_to = _query_value(params, "created_to")
    operator = _query_value(params, "operator")
    if created_from:
        queryset = queryset.filter(created_at__gte=created_from)
    if created_to:
        queryset = queryset.filter(created_at__lte=created_to)
    if operator:
        operator_filter = Q(actor__email__icontains=operator)
        try:
            operator_filter |= Q(actor_id=uuid.UUID(operator))
        except (TypeError, ValueError):
            pass
        queryset = queryset.filter(operator_filter)
    return queryset


def _event_business_context(event: AuditEvent) -> dict[str, Any]:
    context = event.context or {}
    nested = context.get("context", {})
    return nested if isinstance(nested, dict) else {}


def _json_text(value: Any) -> str:
    return json.dumps(value or {}, ensure_ascii=False, sort_keys=True, default=str)


def _repair_event_row(event: AuditEvent) -> dict[str, str]:
    context = _event_business_context(event)
    results = context.get("results", [])
    action_codes = []
    if isinstance(results, list):
        action_codes = [str(item.get("action_code", "")) for item in results if isinstance(item, dict)]
    return {
        "repair_id": str(event.id),
        "operator_id": str(event.actor_id or ""),
        "operator_email": getattr(event.actor, "email", "") if event.actor_id else "",
        "timestamp": event.created_at.isoformat() if event.created_at else "",
        "action": str((event.context or {}).get("action") or "payouts.repair_execution"),
        "status": str((event.context or {}).get("status") or ""),
        "batch_size": str(context.get("batch_size", "")),
        "repaired_count": str(context.get("repaired_count", "")),
        "skipped_count": str(context.get("skipped_count", "")),
        "manual_review_count": str(context.get("manual_review_count", "")),
        "actions": ", ".join(filter(None, action_codes)),
        "result": _json_text(results),
        "filters": _json_text(context.get("filters", {})),
    }


def _repair_action_rows(event: AuditEvent) -> list[dict[str, str]]:
    context = _event_business_context(event)
    results = context.get("results", [])
    if not isinstance(results, list):
        return []
    rows = []
    for result in results:
        if not isinstance(result, dict):
            continue
        rows.append(
            {
                "repair_id": str(event.id),
                "timestamp": event.created_at.isoformat() if event.created_at else "",
                "operator_email": getattr(event.actor, "email", "") if event.actor_id else "",
                "issue_code": str(result.get("issue_code") or ""),
                "action_code": str(result.get("action_code") or ""),
                "status": str(result.get("status") or ""),
                "payout_id": str(result.get("payout_id") or ""),
                "wallet_id": str(result.get("wallet_id") or ""),
                "ledger_entry_id": str(result.get("ledger_entry_id") or ""),
                "amount": str(result.get("amount") or ""),
                "currency": str(result.get("currency") or ""),
                "reason": str(result.get("reason") or ""),
            }
        )
    return rows


def _repair_audit_filters(params) -> dict[str, str]:
    return {key: _query_value(params, key) for key in ["operator", "created_from", "created_to"]}


def _audit_repair_audit_export(*, request, export_format: str, filename: str, filters: dict[str, str], exported_rows: int, total_rows: int, limit: int) -> None:
    AuditService.log_admin_action(
        request=request,
        action="payouts.repair_audit_export",
        target_type="payout_repair_audit_export",
        target_id=export_format,
        context={
            "export_type": "repair_audit",
            "format": export_format,
            "filename": filename,
            "filters": filters,
            "exported_rows": exported_rows,
            "total_rows": total_rows,
            "limit": limit,
            "truncated": total_rows > exported_rows,
        },
    )


class AdminPayoutOpsSummaryAPIView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        return Response(build_payout_admin_ops_summary(request.query_params))


class AdminPayoutOpsReconciliationSnapshotAPIView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        report = PayoutService.build_reconciliation_report()
        safe_report = _json_safe(report)
        return Response(
            {
                "generated_at": datetime.now(UTC).isoformat().replace("+00:00", "Z"),
                "mode": "read_only_snapshot",
                "summary": {
                    "status": safe_report.get("status", "unknown"),
                    "issue_count": safe_report.get("issue_count", 0),
                    "checked_at": safe_report.get("checked_at", ""),
                },
                "snapshot": safe_report,
                "actions": {
                    "repair_performed": False,
                    "note": "This endpoint is read-only. It does not mutate payout requests, wallets or ledger entries.",
                },
            }
        )


class AdminPayoutOpsIntegrityAPIView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        return Response(build_payout_integrity_snapshot(request.query_params))


class AdminPayoutOpsRepairPreviewAPIView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        return Response(build_payout_repair_preview(request.query_params))

    def post(self, request):
        return Response(build_payout_repair_preview(request.data))


class AdminPayoutOpsReconciliationReportExportCsvAPIView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        limit = _positive_int(_query_value(request.query_params, "limit"), default=EXPORT_LIMIT, maximum=EXPORT_LIMIT)
        dataset = _build_reconciliation_export_dataset(request.query_params, limit=limit)
        filename = "payout_reconciliation_report_export.csv"
        rows = []
        for section_rows in dataset["sections"].values():
            rows.extend([[row.get(column, "") for column in RECONCILIATION_EXPORT_HEADER] for row in section_rows])
        _audit_reconciliation_export(
            request=request,
            export_format="csv",
            filename=filename,
            filters=dataset["filters"],
            counts=dataset["counts"],
            limit=limit,
        )
        return _csv_response(filename=filename, header=RECONCILIATION_EXPORT_HEADER, rows=rows)


class AdminPayoutOpsReconciliationReportExportXlsxAPIView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        limit = _positive_int(_query_value(request.query_params, "limit"), default=EXPORT_LIMIT, maximum=EXPORT_LIMIT)
        dataset = _build_reconciliation_export_dataset(request.query_params, limit=limit)
        filename = "payout_reconciliation_report_export.xlsx"
        workbook = Workbook()
        first = True
        for sheet_name, section_rows in dataset["sections"].items():
            worksheet = workbook.active if first else workbook.create_sheet()
            first = False
            worksheet.title = sheet_name[:31]
            worksheet.append(RECONCILIATION_EXPORT_HEADER)
            for row in section_rows:
                worksheet.append(_xlsx_safe_row(row.get(column, "") for column in RECONCILIATION_EXPORT_HEADER))
        _audit_reconciliation_export(
            request=request,
            export_format="xlsx",
            filename=filename,
            filters=dataset["filters"],
            counts=dataset["counts"],
            limit=limit,
        )
        return _xlsx_response(filename=filename, workbook=workbook)


class AdminPayoutOpsRepairAuditExportCsvAPIView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        limit = _positive_int(_query_value(request.query_params, "limit"), default=EXPORT_LIMIT, maximum=EXPORT_LIMIT)
        queryset = _repair_audit_queryset(request.query_params)
        total_rows = queryset.count()
        events = list(queryset[:limit])
        filename = "payout_repair_audit_export.csv"
        rows = [[row.get(column, "") for column in REPAIR_AUDIT_EXPORT_HEADER] for row in [_repair_event_row(event) for event in events]]
        _audit_repair_audit_export(
            request=request,
            export_format="csv",
            filename=filename,
            filters=_repair_audit_filters(request.query_params),
            exported_rows=len(rows),
            total_rows=total_rows,
            limit=limit,
        )
        return _csv_response(filename=filename, header=REPAIR_AUDIT_EXPORT_HEADER, rows=rows)


class AdminPayoutOpsRepairAuditExportXlsxAPIView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        limit = _positive_int(_query_value(request.query_params, "limit"), default=EXPORT_LIMIT, maximum=EXPORT_LIMIT)
        queryset = _repair_audit_queryset(request.query_params)
        total_rows = queryset.count()
        events = list(queryset[:limit])
        filename = "payout_repair_audit_export.xlsx"
        workbook = Workbook()
        repairs_sheet = workbook.active
        repairs_sheet.title = "repairs"
        repairs_sheet.append(REPAIR_AUDIT_EXPORT_HEADER)
        for row in [_repair_event_row(event) for event in events]:
            repairs_sheet.append(_xlsx_safe_row(row.get(column, "") for column in REPAIR_AUDIT_EXPORT_HEADER))

        actions_sheet = workbook.create_sheet("actions")
        actions_sheet.append(REPAIR_AUDIT_ACTION_HEADER)
        for event in events:
            for row in _repair_action_rows(event):
                actions_sheet.append(_xlsx_safe_row(row.get(column, "") for column in REPAIR_AUDIT_ACTION_HEADER))

        _audit_repair_audit_export(
            request=request,
            export_format="xlsx",
            filename=filename,
            filters=_repair_audit_filters(request.query_params),
            exported_rows=len(events),
            total_rows=total_rows,
            limit=limit,
        )
        return _xlsx_response(filename=filename, workbook=workbook)


class AdminPayoutOpsRequestsExportAPIView(APIView):
    permission_classes = [IsAdminUser]

    def get_queryset(self, request):
        queryset = PayoutRequest.objects.select_related("trainer", "trainer__user", "wallet").order_by("-created_at")
        params = request.query_params

        status = _query_value(params, "status")
        trainer_id = _query_value(params, "trainer_id")
        currency = _query_value(params, "currency")

        if status:
            if status == PayoutRequest.Status.PENDING:
                queryset = queryset.filter(status__in=[PayoutRequest.Status.PENDING, PayoutRequest.Status.REQUESTED])
            else:
                queryset = queryset.filter(status=status)
        if trainer_id:
            queryset = queryset.filter(Q(trainer__user_id=trainer_id) | Q(trainer_id=trainer_id))
        if currency:
            queryset = queryset.filter(currency=currency.upper())
        return _apply_created_filters(queryset, params)

    def get(self, request):
        limit = _positive_int(_query_value(request.query_params, "limit"), default=EXPORT_LIMIT, maximum=EXPORT_LIMIT)
        queryset = self.get_queryset(request)
        total_rows = queryset.count()
        payouts = list(queryset[:limit])
        filename = "payout_admin_requests_export.csv"

        rows = [
            [
                str(payout.id),
                str(payout.trainer_id),
                getattr(payout.trainer, "display_name", ""),
                payout.status,
                _money(payout.amount),
                payout.currency,
                payout.destination_masked,
                payout.created_at.isoformat(),
                payout.updated_at.isoformat(),
            ]
            for payout in payouts
        ]

        filters = {key: _query_value(request.query_params, key) for key in ["status", "trainer_id", "currency", "created_from", "created_to"]}
        _audit_context(
            request=request,
            export_type="requests",
            filename=filename,
            filters=filters,
            exported_rows=len(rows),
            total_rows=total_rows,
            limit=limit,
        )

        return _csv_response(
            filename=filename,
            header=["id", "trainer_id", "trainer_name", "status", "amount", "currency", "destination_masked", "created_at", "updated_at"],
            rows=rows,
        )


class AdminPayoutOpsLedgerExportAPIView(APIView):
    permission_classes = [IsAdminUser]

    def get_queryset(self, request):
        queryset = (
            BalanceEntry.objects.select_related("wallet", "wallet__trainer", "wallet__trainer__user")
            .order_by("-created_at")
        )
        params = request.query_params

        status = _query_value(params, "status")
        trainer_id = _query_value(params, "trainer_id")
        currency = _query_value(params, "currency")
        entry_type = _query_value(params, "entry_type")
        direction = _query_value(params, "direction")
        source_type = _query_value(params, "source_type")

        if status:
            queryset = queryset.filter(status=status)
        if trainer_id:
            queryset = queryset.filter(Q(wallet__trainer__user_id=trainer_id) | Q(wallet__trainer_id=trainer_id))
        if currency:
            queryset = queryset.filter(currency=currency.upper())
        if entry_type:
            queryset = queryset.filter(entry_type=entry_type)
        if direction:
            queryset = queryset.filter(direction=direction)
        if source_type:
            queryset = queryset.filter(source_type=source_type)
        return _apply_created_filters(queryset, params)

    def get(self, request):
        limit = _positive_int(_query_value(request.query_params, "limit"), default=EXPORT_LIMIT, maximum=EXPORT_LIMIT)
        queryset = self.get_queryset(request)
        total_rows = queryset.count()
        entries = list(queryset[:limit])
        filename = "payout_admin_ledger_export.csv"

        rows = [
            [
                str(entry.id),
                str(entry.wallet.trainer.user_id),
                getattr(entry.wallet.trainer, "display_name", ""),
                entry.entry_type,
                entry.direction,
                entry.status,
                _money(entry.amount),
                entry.currency,
                entry.source_type,
                str(entry.source_id),
                entry.created_at.isoformat(),
                entry.updated_at.isoformat(),
            ]
            for entry in entries
        ]

        filters = {
            key: _query_value(request.query_params, key)
            for key in ["status", "trainer_id", "currency", "entry_type", "direction", "source_type", "created_from", "created_to"]
        }
        _audit_context(
            request=request,
            export_type="ledger",
            filename=filename,
            filters=filters,
            exported_rows=len(rows),
            total_rows=total_rows,
            limit=limit,
        )

        return _csv_response(
            filename=filename,
            header=[
                "id",
                "trainer_id",
                "trainer_name",
                "entry_type",
                "direction",
                "status",
                "amount",
                "currency",
                "source_type",
                "source_id",
                "created_at",
                "updated_at",
            ],
            rows=rows,
        )
