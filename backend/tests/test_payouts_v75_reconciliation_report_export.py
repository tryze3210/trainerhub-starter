from __future__ import annotations

import csv
import io
from decimal import Decimal

import pytest
from django.contrib.auth import get_user_model
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.audit.models import AuditEvent
from apps.payouts.models import BalanceEntry, PayoutRequest, TrainerWallet
from apps.trainers.models import TrainerProfile


@pytest.fixture
def api_client():
    return APIClient()


@pytest.fixture
def admin_user(db):
    User = get_user_model()
    return User.objects.create_superuser(email="admin-payout-v75@example.com", password="strong-pass-123")


@pytest.fixture
def regular_user(db):
    User = get_user_model()
    return User.objects.create_user(email="regular-payout-v75@example.com", password="strong-pass-123")


@pytest.fixture
def reconciliation_export_dataset(db):
    User = get_user_model()
    trainer_user = User.objects.create_user(
        email="trainer-payout-v75@example.com",
        password="strong-pass-123",
        role="trainer",
    )
    trainer = TrainerProfile.objects.create(
        user=trainer_user,
        slug="trainer-payout-v75",
        display_name="Trainer Payout V75",
        status="active",
    )
    wallet = TrainerWallet.objects.create(
        trainer=trainer,
        currency="RUB",
        available_amount=Decimal("800.00"),
        pending_amount=Decimal("0.00"),
        locked_amount=Decimal("250.00"),
    )
    payout = PayoutRequest.objects.create(
        trainer=trainer,
        wallet=wallet,
        amount=Decimal("250.00"),
        currency="RUB",
        status=PayoutRequest.Status.PENDING,
        destination_json={"destination_masked": "**** 7575"},
    )
    reserve = BalanceEntry.objects.create(
        wallet=wallet,
        entry_type=BalanceEntry.EntryType.RESERVE,
        direction="debit",
        amount=Decimal("250.00"),
        currency="RUB",
        status="locked",
        source_type="payout_request",
        source_id=payout.id,
    )
    release = BalanceEntry.objects.create(
        wallet=wallet,
        entry_type=BalanceEntry.EntryType.RELEASE,
        direction="credit",
        amount=Decimal("25.00"),
        currency="RUB",
        status="released",
        source_type="payout_request",
        source_id=payout.id,
    )
    return {"trainer": trainer, "wallet": wallet, "payout": payout, "reserve": reserve, "release": release}


def _read_csv(response):
    body = response.content.decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(body)))


@pytest.mark.django_db
def test_admin_can_export_payout_reconciliation_report_csv(api_client, admin_user, reconciliation_export_dataset):
    api_client.force_authenticate(admin_user)

    response = api_client.get("/api/v1/payouts/admin-ops/reconciliation/export.csv?currency=RUB&limit=20")

    assert response.status_code == 200
    assert response["Content-Type"].startswith("text/csv")
    assert "payout_reconciliation_report_export.csv" in response["Content-Disposition"]
    rows = _read_csv(response)
    sections = {row["section"] for row in rows}
    assert {"payout", "reserve", "release", "balance", "integrity_status"}.issubset(sections)
    assert any(row["section"] == "payout" and row["payout_id"] == str(reconciliation_export_dataset["payout"].id) for row in rows)
    assert any(row["section"] == "reserve" and row["source_id"] == str(reconciliation_export_dataset["payout"].id) for row in rows)
    assert any(row["section"] == "release" and row["amount"] == "25.00" for row in rows)
    assert any(row["section"] == "balance" and row["available_amount"] == "800.00" for row in rows)
    assert any(row["section"] == "integrity_status" and row["integrity_status"] in {"healthy", "attention_required"} for row in rows)

    event = AuditEvent.objects.get(
        event_type="admin.payouts.reconciliation_report_export",
        entity_type="payout_reconciliation_export",
        entity_id="csv",
    )
    assert event.actor == admin_user
    assert event.context["context"]["format"] == "csv"
    assert event.context["context"]["section_counts"]["payouts"] == 1
    assert event.context["context"]["exported_rows"] == len(rows)


@pytest.mark.django_db
def test_admin_can_export_payout_reconciliation_report_xlsx(api_client, admin_user, reconciliation_export_dataset):
    api_client.force_authenticate(admin_user)

    response = api_client.get("/api/v1/payouts/admin-ops/reconciliation/export.xlsx?currency=RUB&limit=20")

    assert response.status_code == 200
    assert response["Content-Type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert "payout_reconciliation_report_export.xlsx" in response["Content-Disposition"]

    workbook = load_workbook(io.BytesIO(response.content))
    assert {"payouts", "reserves", "releases", "balances", "integrity_status"}.issubset(set(workbook.sheetnames))
    payout_sheet = workbook["payouts"]
    headers = [cell.value for cell in payout_sheet[1]]
    payout_id_index = headers.index("payout_id") + 1
    assert payout_sheet.cell(row=2, column=payout_id_index).value == str(reconciliation_export_dataset["payout"].id)

    event = AuditEvent.objects.get(
        event_type="admin.payouts.reconciliation_report_export",
        entity_type="payout_reconciliation_export",
        entity_id="xlsx",
    )
    assert event.context["context"]["format"] == "xlsx"
    assert event.context["context"]["section_counts"]["balances"] == 1


@pytest.mark.django_db
def test_non_admin_cannot_export_payout_reconciliation_report(api_client, regular_user, reconciliation_export_dataset):
    api_client.force_authenticate(regular_user)

    response = api_client.get("/api/v1/payouts/admin-ops/reconciliation/export.csv")

    assert response.status_code == 403
