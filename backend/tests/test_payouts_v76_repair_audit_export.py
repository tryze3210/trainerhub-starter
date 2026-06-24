from __future__ import annotations

import csv
import io

import pytest
from django.contrib.auth import get_user_model
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.audit.models import AuditEvent


@pytest.fixture
def api_client():
    return APIClient()


@pytest.fixture
def admin_user(db):
    User = get_user_model()
    return User.objects.create_superuser(email="admin-payout-v76@example.com", password="strong-pass-123")


@pytest.fixture
def operator_user(db):
    User = get_user_model()
    return User.objects.create_user(email="operator-payout-v76@example.com", password="strong-pass-123")


@pytest.fixture
def regular_user(db):
    User = get_user_model()
    return User.objects.create_user(email="regular-payout-v76@example.com", password="strong-pass-123")


@pytest.fixture
def repair_audit_event(db, operator_user):
    return AuditEvent.objects.create(
        actor=operator_user,
        event_type="admin.payouts.repair_execution",
        entity_type="payout_repair_execution",
        entity_id="v73",
        context={
            "action": "payouts.repair_execution",
            "status": "completed",
            "context": {
                "filters": {"currency": "RUB", "status": "pending"},
                "batch_size": 10,
                "repaired_count": 1,
                "skipped_count": 0,
                "manual_review_count": 1,
                "results": [
                    {
                        "issue_code": "locked_balance_mismatch",
                        "action_code": "release_excess_locked_to_available",
                        "status": "repaired",
                        "wallet_id": "wallet-v76",
                        "ledger_entry_id": "ledger-v76",
                        "amount": "50.00",
                        "currency": "RUB",
                    },
                    {
                        "issue_code": "paid_payout_missing_payout_ledger",
                        "action_code": "manual_review_required",
                        "status": "manual_review_required",
                        "payout_id": "payout-v76",
                        "reason": "Paid payout is missing payout ledger evidence.",
                        "currency": "RUB",
                    },
                ],
            },
        },
    )


def _read_csv(response):
    body = response.content.decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(body)))


@pytest.mark.django_db
def test_admin_can_export_payout_repair_audit_csv(api_client, admin_user, repair_audit_event):
    api_client.force_authenticate(admin_user)

    response = api_client.get("/api/v1/payouts/admin-ops/repair/audit/export.csv?operator=operator-payout-v76&limit=20")

    assert response.status_code == 200
    assert response["Content-Type"].startswith("text/csv")
    assert "payout_repair_audit_export.csv" in response["Content-Disposition"]
    rows = _read_csv(response)
    assert len(rows) == 1
    assert rows[0]["repair_id"] == str(repair_audit_event.id)
    assert rows[0]["operator_email"] == "operator-payout-v76@example.com"
    assert rows[0]["repaired_count"] == "1"
    assert rows[0]["manual_review_count"] == "1"
    assert "release_excess_locked_to_available" in rows[0]["actions"]
    assert "paid_payout_missing_payout_ledger" in rows[0]["result"]

    export_event = AuditEvent.objects.get(
        event_type="admin.payouts.repair_audit_export",
        entity_type="payout_repair_audit_export",
        entity_id="csv",
    )
    assert export_event.actor == admin_user
    assert export_event.context["context"]["format"] == "csv"
    assert export_event.context["context"]["exported_rows"] == 1
    assert export_event.context["context"]["total_rows"] == 1


@pytest.mark.django_db
def test_admin_can_export_payout_repair_audit_xlsx(api_client, admin_user, repair_audit_event):
    api_client.force_authenticate(admin_user)

    response = api_client.get("/api/v1/payouts/admin-ops/repair/audit/export.xlsx?limit=20")

    assert response.status_code == 200
    assert response["Content-Type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert "payout_repair_audit_export.xlsx" in response["Content-Disposition"]

    workbook = load_workbook(io.BytesIO(response.content))
    assert {"repairs", "actions"}.issubset(set(workbook.sheetnames))
    repairs_headers = [cell.value for cell in workbook["repairs"][1]]
    operator_index = repairs_headers.index("operator_email") + 1
    assert workbook["repairs"].cell(row=2, column=operator_index).value == "operator-payout-v76@example.com"

    action_headers = [cell.value for cell in workbook["actions"][1]]
    issue_index = action_headers.index("issue_code") + 1
    action_values = [workbook["actions"].cell(row=row, column=issue_index).value for row in range(2, 4)]
    assert "locked_balance_mismatch" in action_values
    assert "paid_payout_missing_payout_ledger" in action_values

    export_event = AuditEvent.objects.get(
        event_type="admin.payouts.repair_audit_export",
        entity_type="payout_repair_audit_export",
        entity_id="xlsx",
    )
    assert export_event.context["context"]["format"] == "xlsx"


@pytest.mark.django_db
def test_non_admin_cannot_export_payout_repair_audit(api_client, regular_user, repair_audit_event):
    api_client.force_authenticate(regular_user)

    response = api_client.get("/api/v1/payouts/admin-ops/repair/audit/export.csv")

    assert response.status_code == 403
